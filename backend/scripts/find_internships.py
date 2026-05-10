import asyncio
import hashlib
import os
import re
import sys
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Set
from urllib.parse import urlparse, urljoin

import httpx
import whois
# Fix for Windows console encoding issues with Rich
if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

from bs4 import BeautifulSoup
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from rich.console import Console
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn
from rich.table import Table
from rich.panel import Panel

# Load environment variables
load_dotenv()

# Configuration
ADZUNA_APP_ID = os.getenv("ADZUNA_APP_ID")
ADZUNA_API_KEY = os.getenv("ADZUNA_API_KEY")
GOOGLE_MAPS_API_KEY = os.getenv("GOOGLE_MAPS_API_KEY")
HTTP_TIMEOUT = int(os.getenv("HTTP_TIMEOUT", 8))
MAX_RESULTS_PER_SOURCE = int(os.getenv("MAX_RESULTS_PER_SOURCE", 50))

TARGET_DOMAINS = [
    "Android App Development",
    "Game Development",
    "iOS App Development",
    "Web Development",
    "Graphic Design",
    "Data Science"
]

DOMAIN_TO_TAB = {
    "Android App Development": "APP ANDROID",
    "Game Development": "Game development",
    "iOS App Development": "APP IOS",
    "Web Development": "WEB DEVELOPMENT",
    "Graphic Design": "Graphic Design",
    "Data Science": "DATA SCIENCE"
}

KNOWN_ATS_DOMAINS = {
    "internshala.com", "linkedin.com", "naukri.com", "unstop.com",
    "wellfound.com", "lever.co", "greenhouse.io", "workday.com",
    "smartrecruiters.com", "ashbyhq.com", "workindia.in"
}

FRAUD_KEYWORDS = [
    "training fee", "registration fee", "security deposit",
    "certificate fee", "joining fee", "pay to work",
    "pay to apply", "investment required", "refundable deposit",
    "refer and earn", "mlm", "downline", "earn from home"
]

USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"

console = Console()

class Internship:
    def __init__(self, domain: str, source: str):
        self.domain = domain
        self.source = source
        self.company_name = ""
        self.title = ""
        self.stipend_display = "Not mentioned"
        self.stipend_min = 0
        self.duration = "Not mentioned"
        self.role = ""
        self.last_date: Optional[datetime] = None
        self.apply_link = ""
        self.location = "Remote"
        self.status = "PENDING"
        self.company_status = "UNVERIFIED"
        self.rejection_reason = ""
        self.trust_score = 0
        self.created_at = datetime.now()
        self.checked_at = datetime.now()

    def get_hash(self) -> str:
        data = f"{self.company_name.lower()}{self.title.lower()}{self.domain}"
        return hashlib.sha256(data.encode()).hexdigest()

class InternshipFinder:
    def __init__(self, dry_run=False):
        self.dry_run = dry_run
        self.seen_hashes: Set[str] = set()
        self.results: List[Internship] = []
        self.stats = {
            "total_found": 0,
            "active": 0,
            "rejected": 0,
            "expired": 0,
            "needs_review": 0
        }

    async def fetch_adzuna(self, domain: str) -> List[Internship]:
        console.print(f"[cyan]Adzuna APP_ID loaded: {bool(ADZUNA_APP_ID)} | KEY loaded: {bool(ADZUNA_API_KEY)}[/cyan]")
        if not ADZUNA_APP_ID or not ADZUNA_API_KEY:
            console.print("[yellow]Adzuna credentials missing, skipping Source 1[/yellow]")
            return []

        keywords = {
            "Android App Development": "android developer intern",
            "Game Development": "game developer intern unity unreal",
            "iOS App Development": "ios developer intern swift",
            "Web Development": "web developer intern frontend backend",
            "Graphic Design": "graphic design intern ui designer",
            "Data Science": "data science intern machine learning"
        }

        url = "https://api.adzuna.com/v1/api/jobs/in/search/1"
        params = {
            "app_id": ADZUNA_APP_ID,
            "app_key": ADZUNA_API_KEY,
            "results_per_page": MAX_RESULTS_PER_SOURCE,
            "what": keywords.get(domain, "intern"),
            "content-type": "application/json"
        }

        found = []
        try:
            async with httpx.AsyncClient(timeout=HTTP_TIMEOUT) as client:
                response = await client.get(url, params=params)
                console.print(f"[cyan]Adzuna status: {response.status_code}[/cyan]")
                if response.status_code != 200:
                    console.print(f"[red]Adzuna response: {response.text[:300]}[/red]")
                if response.status_code == 200:
                    data = response.json()
                    for job in data.get("results", []):
                        intern = Internship(domain, "Adzuna")
                        intern.company_name = job.get("company", {}).get("display_name", "Unknown")
                        intern.title = job.get("title", "")
                        intern.apply_link = job.get("redirect_url", "")
                        intern.location = job.get("location", {}).get("display_name", "Remote")
                        salary_min = job.get("salary_min", 0)
                        salary_max = job.get("salary_max", 0)
                        if salary_min > 0:
                            intern.stipend_min = int(salary_min)
                            if salary_max > salary_min:
                                intern.stipend_display = f"₹{int(salary_min):,} - ₹{int(salary_max):,}/month"
                            else:
                                intern.stipend_display = f"₹{int(salary_min):,}/month"
                        else:
                            intern.stipend_display = "Not mentioned"
                            intern.stipend_min = 0
                        found.append(intern)
        except Exception as e:
            console.print(f"[red]Adzuna error: {str(e)}[/red]")
        
        return found

    async def fetch_internshala(self, domain: str) -> List[Internship]:
        keywords = {
            "Android App Development": "android-development",
            "Game Development": "game-development",
            "iOS App Development": "ios-development",
            "Web Development": "web-development",
            "Graphic Design": "graphic-designing",
            "Data Science": "data-science"
        }
        slug = keywords.get(domain)
        if not slug: return []

        url = f"https://internshala.com/internships/{slug}-internship/"
        found = []
        try:
            async with httpx.AsyncClient(
                headers={
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
                    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                    "Accept-Language": "en-US,en;q=0.5",
                },
                timeout=15,
                follow_redirects=True
            ) as client:
                response = await client.get(url)
                console.print(f"[cyan]Internshala status: {response.status_code} for {slug}[/cyan]")
                
                if response.status_code != 200:
                    console.print(f"[red]Internshala blocked: {response.status_code}[/red]")
                    return []

                soup = BeautifulSoup(response.text, 'html.parser')
                
                # Try multiple possible container selectors
                containers = (
                    soup.find_all('div', class_='individual_internship') or
                    soup.find_all('div', attrs={'data-internship_id': True}) or
                    soup.find_all('div', class_=re.compile(r'internship.*card|card.*internship', re.I))
                )
                
                console.print(f"[cyan]Internshala containers found: {len(containers)}[/cyan]")
                
                if not containers:
                    # Save HTML snippet for debugging
                    console.print(f"[yellow]Internshala HTML preview: {response.text[2000:3000]}[/yellow]")
                    return []

                for container in containers[:MAX_RESULTS_PER_SOURCE]:
                    try:
                        intern = Internship(domain, "Internshala")
                        
                        # Title — try multiple selectors
                        title_el = (
                            container.find('h3', class_='job-internship-name') or
                            container.find('h3') or
                            container.find('a', class_=re.compile(r'job|intern|title', re.I))
                        )
                        if not title_el: continue
                        intern.title = title_el.get_text(strip=True)
                        
                        # Company
                        company_el = (
                            container.find('div', class_='company-name') or
                            container.find('p', class_='company-name') or
                            container.find(class_=re.compile(r'company', re.I))
                        )
                        intern.company_name = company_el.get_text(strip=True) if company_el else "Unknown"
                        
                        # Stipend
                        stipend_el = (
                            container.find('span', class_='stipend') or
                            container.find(class_=re.compile(r'stipend|salary', re.I))
                        )
                        if stipend_el:
                            intern.stipend_display = stipend_el.get_text(strip=True)
                        
                        # Location
                        loc_el = (
                            container.find('div', class_='location_link') or
                            container.find(class_=re.compile(r'location', re.I))
                        )
                        if loc_el:
                            intern.location = loc_el.get_text(strip=True)
                        
                        # Duration
                        dur_el = container.find(class_=re.compile(r'duration', re.I))
                        if dur_el:
                            intern.duration = dur_el.get_text(strip=True)
                        
                        # Apply link
                        link_el = (
                            container.find('a', class_='view_detail_button') or
                            container.find('a', href=re.compile(r'/internships/detail|/internship/'))
                        )
                        if link_el and link_el.get('href'):
                            intern.apply_link = urljoin("https://internshala.com", link_el['href'])
                        
                        if intern.title and intern.apply_link:
                            found.append(intern)
                            
                    except Exception as e:
                        continue
                        
        except Exception as e:
            console.print(f"[red]Internshala error: {str(e)}[/red]")
        
        console.print(f"[green]Internshala found: {len(found)} for {domain}[/green]")
        return found

    async def fetch_naukri(self, domain: str) -> List[Internship]:
        keywords = {
            "Android App Development": "android+developer+internship",
            "Game Development": "game+developer+internship",
            "iOS App Development": "ios+developer+internship",
            "Web Development": "web+developer+internship",
            "Graphic Design": "graphic+designer+internship",
            "Data Science": "data+science+internship"
        }
        kw = keywords.get(domain, "internship")
        url = f"https://www.naukri.com/{kw.replace('+','-')}-jobs"
        found = []
        try:
            async with httpx.AsyncClient(
                headers={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"},
                timeout=12,
                follow_redirects=True
            ) as client:
                resp = await client.get(url)
                if resp.status_code == 200:
                    soup = BeautifulSoup(resp.text, 'html.parser')
                    # Broader selectors for Naukri
                    cards = (
                        soup.find_all('article', class_=re.compile(r'jobTuple|job-card', re.I)) or
                        soup.find_all('div', class_=re.compile(r'jobTuple|job-card|cust-job-tuple', re.I)) or
                        soup.find_all(attrs={"data-job-id": True})
                    )
                    console.print(f"[cyan]Naukri cards: {len(cards)} for {domain}[/cyan]")
                    for card in cards[:MAX_RESULTS_PER_SOURCE]:
                        try:
                            intern = Internship(domain, "Naukri")
                            title_el = card.find('a', class_=re.compile(r'title|jobTitle', re.I))
                            company_el = card.find(class_=re.compile(r'companyInfo|company-name', re.I))
                            salary_el = card.find(class_=re.compile(r'salary|stipend', re.I))
                            loc_el = card.find(class_=re.compile(r'location|loc', re.I))
                            
                            if not title_el: continue
                            intern.title = title_el.get_text(strip=True)
                            intern.apply_link = title_el.get('href', '')
                            if not intern.apply_link.startswith('http'):
                                intern.apply_link = 'https://www.naukri.com' + intern.apply_link
                            intern.company_name = company_el.get_text(strip=True) if company_el else "Unknown"
                            intern.stipend_display = salary_el.get_text(strip=True) if salary_el else "Not mentioned"
                            intern.location = loc_el.get_text(strip=True) if loc_el else "Not mentioned"
                            
                            if intern.title and intern.apply_link:
                                found.append(intern)
                        except:
                            continue
        except Exception as e:
            console.print(f"[red]Naukri error: {str(e)}[/red]")
        return found


    def normalize_internship(self, intern: Internship):
        # 1. Parse stipend
        text = intern.stipend_display.replace(',', '').replace('₹', '').replace('\u20b9', '').lower().strip()
        numbers = re.findall(r'\d+', text)
        if numbers:
            val = int(numbers[0])
            if val < 100:  # only multiply if clearly "k" shorthand
                val *= 1000
            intern.stipend_min = val
        elif "interview" in text or "performance" in text:
            intern.stipend_min = 1  # treat as paid (medium confidence)
        else:
            intern.stipend_min = 0
        
        # 2. Location
        loc = intern.location.lower()
        if any(x in loc for x in ["wfh", "remote", "work from home"]):
            intern.location = "Work from home"
        elif any(x in loc for x in ["onsite", "on-site", "in office"]):
            intern.location = "On-site"
            
        # 3. Role
        clean_role = intern.title
        clean_role = re.sub(r'\s*-\s*Internship', ' Intern', clean_role, flags=re.I)
        clean_role = re.sub(r'\(.*?\)', '', clean_role).strip()
        intern.role = clean_role

        # 4. Last Date (generic parser)
        # Note: Scrapers should ideally provide this. If not, set to None.
        pass

    async def run_pipeline(self, intern: Internship):
        # STEP 1: NORMALIZE
        self.normalize_internship(intern)

        # STEP 2: PAID GATE
        is_interview_basis = "interview basis" in intern.stipend_display.lower()
        if intern.stipend_min <= 0 and not is_interview_basis:
            intern.status = "REJECTED"
            intern.rejection_reason = "unpaid_or_missing_stipend"
            return

        # STEP 3: FRAUD DETECTION
        combined_text = (intern.title + " " + intern.stipend_display).lower()
        for kw in FRAUD_KEYWORDS:
            if kw in combined_text:
                intern.status = "REJECTED"
                intern.rejection_reason = f"fraud_keyword:{kw}"
                return

        # STEP 4: LINK VALIDATION
        if any(ats in intern.apply_link.lower() for ats in KNOWN_ATS_DOMAINS):
            # ATS links are pre-validated — skip HTTP check
            pass
        else:
            try:
                async with httpx.AsyncClient(follow_redirects=True, timeout=8) as client:
                    resp = await client.get(intern.apply_link)
                    final_url = str(resp.url)
                    domain = urlparse(final_url).netloc.lower()
                    
                    if resp.status_code >= 400:
                        intern.status = "REJECTED"
                        intern.rejection_reason = f"invalid_link:HTTP_{resp.status_code}"
                        return
                    
                    if any(x in domain for x in ["wa.me", "t.me", "bit.ly"]):
                        intern.status = "REJECTED"
                        intern.rejection_reason = "invalid_link:suspicious_domain"
                        return
            except Exception as e:
                intern.status = "REJECTED"
                intern.rejection_reason = f"invalid_link:timeout_or_error"
                return

        # STEP 5: COMPANY VERIFICATION
        intern.trust_score = await self.calculate_trust_score(intern)
        if intern.trust_score >= 2:
            intern.company_status = "VERIFIED"
        elif intern.trust_score == 1:
            intern.company_status = "NEEDS_REVIEW"
        else:
            intern.status = "REJECTED"
            intern.rejection_reason = "low_trust"
            return

        # STEP 6: DEADLINE CHECK
        today = datetime.now().date()
        if intern.last_date and intern.last_date.date() < today:
            intern.status = "EXPIRED"
            intern.rejection_reason = "deadline_passed"
            return
        
        if not intern.last_date:
            # Default expiry check
            if (datetime.now() - intern.created_at).days > 60:
                intern.status = "EXPIRED"
                intern.rejection_reason = "auto_expired_60d"
                return

        # STEP 7: DEDUPLICATION
        h = intern.get_hash()
        if h in self.seen_hashes:
            intern.status = "DUPLICATE"
            return
        self.seen_hashes.add(h)

        if intern.status == "PENDING":
            if intern.company_status == "NEEDS_REVIEW":
                intern.status = "NEEDS_REVIEW"
            else:
                intern.status = "ACTIVE"

    async def calculate_trust_score(self, intern: Internship) -> int:
        if "internshala.com" in intern.apply_link.lower():
            return 3 # Internshala IS the ATS, auto-verified
            
        score = 0
        domain_parts = urlparse(intern.apply_link).netloc.split('.')
        base_domain = '.'.join(domain_parts[-2:]) if len(domain_parts) >= 2 else ""

        # Signal 3: Known ATS (Fast check)
        if any(ats in intern.apply_link.lower() for ats in KNOWN_ATS_DOMAINS):
            score += 1

        # Signal 1 & 2: Website check
        if base_domain and base_domain not in KNOWN_ATS_DOMAINS:
            try:
                async with httpx.AsyncClient(timeout=5) as client:
                    resp = await client.get(f"http://{base_domain}")
                    if resp.status_code < 400:
                        score += 1
                        # Check for real pages
                        if any(x in resp.text.lower() for x in ["/about", "/careers", "/contact"]):
                            score += 1
            except: pass

        # Signal 4: Glassdoor check
        try:
            async with httpx.AsyncClient(
                headers={"User-Agent": USER_AGENT}, timeout=5
            ) as client:
                resp = await client.get(
                    f"https://www.glassdoor.co.in/Search/results.htm?keyword={intern.company_name}"
                )
                if resp.status_code == 200 and intern.company_name.lower()[:6] in resp.text.lower():
                    score += 1
        except:
            pass

        # Signal 5: Domain Age
        if base_domain and base_domain not in KNOWN_ATS_DOMAINS:
            try:
                w = await asyncio.to_thread(whois.whois, base_domain)
                creation_date = w.creation_date
                if isinstance(creation_date, list): creation_date = creation_date[0]
                if creation_date:
                    age_days = (datetime.now() - creation_date).days
                    if age_days >= 365: score += 1
                    elif age_days < 90: score -= 1
            except: pass

        # Signal 6: Internshala company page
        try:
            slug = intern.company_name.lower().replace(" ", "-")
            async with httpx.AsyncClient(headers={"User-Agent": USER_AGENT}, timeout=5) as client:
                resp = await client.get(f"https://internshala.com/companies/{slug}/")
                if resp.status_code == 200: score += 1
        except: pass

        return max(0, score)

    def write_to_excel(self, filename: str):
        wb = Workbook()
        # Remove default sheet
        std = wb["Sheet"]
        wb.remove(std)

        # Create tabs
        tabs: Dict[str, Worksheet] = {}
        all_tab_names = list(DOMAIN_TO_TAB.values()) + ["NEEDS REVIEW", "REJECTED", "EXPIRED"]
        for name in all_tab_names:
            tabs[name] = wb.create_sheet(name)
            self._setup_sheet(tabs[name], name)

        # Counters for Sr No.
        counters = {name: 1 for name in all_tab_names}

        for intern in self.results:
            target_tab = ""
            extra_cols = []

            if intern.status == "REJECTED":
                target_tab = "REJECTED"
                extra_cols = [intern.rejection_reason, intern.checked_at.strftime("%Y-%m-%d %H:%M")]
            elif intern.status == "EXPIRED":
                target_tab = "EXPIRED"
                extra_cols = [intern.checked_at.strftime("%Y-%m-%d %H:%M")]
            elif intern.status == "NEEDS_REVIEW":
                target_tab = "NEEDS REVIEW"
                extra_cols = ["Trust score: " + str(intern.trust_score), intern.checked_at.strftime("%Y-%m-%d %H:%M")]
            elif intern.status == "ACTIVE":
                target_tab = DOMAIN_TO_TAB.get(intern.domain, "REJECTED")
            
            if target_tab:
                ws = tabs[target_tab]
                row_idx = counters[target_tab] + 1
                self._write_row(ws, row_idx, counters[target_tab], intern, extra_cols)
                counters[target_tab] += 1

        # Formatting and Save
        for ws in wb.worksheets:
            self._finalize_sheet(ws)

        os.makedirs(os.path.dirname(filename), exist_ok=True)
        wb.save(filename)

    def _setup_sheet(self, ws: Worksheet, name: str):
        headers = ["Sr. No.", "Company Name", "Stipend", "Name of internship", "Duration", "Role", "Last Date", "Link", "Location"]
        if name == "REJECTED":
            headers += ["Rejection Reason", "Checked At"]
        elif name == "EXPIRED":
            headers += ["Expired At"]
        elif name == "NEEDS REVIEW":
            headers += ["Reason", "Checked At"]
        
        ws.append(headers)
        
        # Header Style
        header_fill = PatternFill(start_color="1A1A1E", end_color="1A1A1E", fill_type="solid")
        header_font = Font(color="F5A623", bold=True, name="Calibri", size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    def _write_row(self, ws: Worksheet, row_idx: int, sr_no: int, intern: Internship, extra: List[str]):
        row_data = [
            sr_no,
            intern.company_name,
            intern.stipend_display,
            intern.title,
            intern.duration,
            intern.role,
            intern.last_date.strftime("%d %b %Y") if intern.last_date else "Not mentioned",
            "Apply",
            intern.location
        ] + extra

        ws.append(row_data)
        
        # Style
        bg_color = "0D0D0F" if row_idx % 2 == 0 else "141417"
        row_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        row_font = Font(color="F8F8F8", name="Calibri", size=10)
        
        for i, cell in enumerate(ws[row_idx]):
            cell.fill = row_fill
            cell.font = row_font
            # Hyperlink for Link column (Index 7)
            if i == 7:
                cell.hyperlink = intern.apply_link
                cell.font = Font(color="F5A623", underline="single", name="Calibri", size=10)

    def _finalize_sheet(self, ws: Worksheet):
        # Column widths
        widths = {'A': 8, 'B': 25, 'C': 22, 'D': 40, 'E': 12, 'F': 28, 'G': 14, 'H': 10, 'I': 18}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        
        # Freeze and Filter
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    async def start(self, domains: List[str], output_file: str):
        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TaskProgressColumn(),
            console=console
        ) as progress:
            
            for domain in domains:
                task_id = progress.add_task(f"Searching {domain}...", total=4)
                
                # Fetch
                results_adzuna = await self.fetch_adzuna(domain)
                progress.update(task_id, advance=1)
                
                results_ishala = await self.fetch_internshala(domain)
                progress.update(task_id, advance=1)
                
                results_naukri = await self.fetch_naukri(domain)
                progress.update(task_id, advance=1)
                
                console.print("[yellow]Unstop skipped — JS-rendered site[/yellow]")
                
                all_found = results_adzuna + results_ishala + results_naukri
                self.stats["total_found"] += len(all_found)
                
                # Pipeline
                item_task = progress.add_task(f"Processing {len(all_found)} items...", total=len(all_found))
                for item in all_found:
                    await self.run_pipeline(item)
                    self.results.append(item)
                    
                    # Update stats
                    if item.status == "ACTIVE": self.stats["active"] += 1
                    elif item.status == "REJECTED": self.stats["rejected"] += 1
                    elif item.status == "EXPIRED": self.stats["expired"] += 1
                    elif item.status == "NEEDS_REVIEW": self.stats["needs_review"] += 1
                    
                    progress.update(item_task, advance=1)
                
                progress.remove_task(item_task)

        self._print_rejection_breakdown()

        if not self.dry_run:
            self.write_to_excel(output_file)
            self._print_summary(output_file)
        else:
            console.print("[bold cyan]Dry run complete. No file written.[/bold cyan]")
            self._print_summary("N/A")

    def _print_summary(self, filename: str):
        table = Table(show_header=False, border_style="bright_blue")
        table.add_row("Total found:", f"{self.stats['total_found']}")
        table.add_row("Active (saved):", f"[green]{self.stats['active']}[/green]")
        table.add_row("Rejected:", f"[red]{self.stats['rejected']}[/red]")
        table.add_row("Expired:", f"[yellow]{self.stats['expired']}[/yellow]")
        table.add_row("Needs Review:", f"[orange1]{self.stats['needs_review']}[/orange1]")
        table.add_row("Output:", f"[bold white]{filename}[/bold white]")

        console.print("\n")
        console.print(Panel(table, title="[bold white]IIB India — Run Complete[/bold white]", expand=False))

    def _print_rejection_breakdown(self):
        from collections import Counter
        reasons = Counter()
        for intern in self.results:
            if intern.status in ("REJECTED", "NEEDS_REVIEW", "EXPIRED", "DUPLICATE"):
                reasons[intern.rejection_reason or intern.status] += 1
            elif intern.status == "ACTIVE":
                reasons["ACTIVE"] += 1
        
        console.print("\n[bold yellow]Rejection Breakdown:[/bold yellow]")
        for reason, count in reasons.most_common():
            color = "green" if reason == "ACTIVE" else "red"
            console.print(f"  [{color}]{reason}: {count}[/{color}]")

def parse_args():
    import argparse
    parser = argparse.ArgumentParser(description="IIB India Internship Finder")
    parser.add_argument("--all", action="store_true", help="Run for all 6 domains")
    parser.add_argument("--domain", type=str, help="Run for a specific domain")
    parser.add_argument("--output", type=str, help="Output filename")
    parser.add_argument("--dry-run", action="store_true", help="Don't save to Excel")
    return parser.parse_args()

async def main():
    args = parse_args()
    
    selected_domains = []
    if args.all:
        selected_domains = TARGET_DOMAINS
    elif args.domain:
        if args.domain in TARGET_DOMAINS:
            selected_domains = [args.domain]
        else:
            console.print(f"[red]Invalid domain. Choose from: {TARGET_DOMAINS}[/red]")
            return
    else:
        # Default to all if nothing specified? User prompt says --all or --domain
        console.print("[yellow]Please specify --all or --domain. See --help.[/yellow]")
        return

    date_str = datetime.now().strftime("%Y_%m_%d")
    default_output = f"e:\\intraget\\iib-india\\output\\IIB_India_Internships_{date_str}.xlsx"
    output_file = args.output if args.output else default_output
    if args.output and not (":" in args.output or args.output.startswith("\\")):
        # If relative path provided for output
        output_file = os.path.join("e:\\intraget\\iib-india\\output\\", args.output)

    finder = InternshipFinder(dry_run=args.dry_run)
    await finder.start(selected_domains, output_file)

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        console.print("\n[red]Process interrupted by user.[/red]")
        sys.exit(1)
    except Exception as e:
        console.print(f"\n[bold red]CRITICAL ERROR: {str(e)}[/bold red]")
        sys.exit(1)
